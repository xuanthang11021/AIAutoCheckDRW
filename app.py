from __future__ import annotations
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
os.environ['http_proxy']  = 'http://150.61.8.70:10080'
os.environ['https_proxy'] = 'http://150.61.8.70:10080'
os.environ["FLAGS_use_onednn"] = "0"
import shutil 
import base64
import io
import sys
import tempfile
import traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
from datetime import datetime
import cv2, gc
import numpy as np
from flask import Flask, jsonify, request, make_response
from flask_cors import CORS


SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from process_drawing import (
    ImageRegionExtractor,
    get_ocr_instance,
    preprocess_image_region,
    read_and_process_text,
    OCR_TYPE,
)
import re
from rsd_validator import RSDValidator

# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------
app = Flask(__name__)
from flask_cors import CORS
CORS(app)

# Khởi tạo Validator (Database version)
validator = RSDValidator()

@app.after_request
def after_request(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response

# Maximum payload size: 50 MB (adjust as needed)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _img_to_base64(img_bgr: np.ndarray) -> str:
    """Encode an OpenCV BGR image to a *data-URI* base64 PNG string."""
    success, buf = cv2.imencode(".png", img_bgr)
    if not success:
        return ""
    return "data:image/png;base64," + base64.b64encode(buf).decode("ascii")


def _img_gray_to_base64(img_gray: np.ndarray) -> str:
    """Encode a single-channel (grayscale / binary) image to base64 PNG."""
    success, buf = cv2.imencode(".png", img_gray)
    if not success:
        return ""
    return "data:image/png;base64," + base64.b64encode(buf).decode("ascii")


def _save_uploaded_image(file_storage) -> str:
    """
    Persist the uploaded image to a temporary file and return its path.
    We need a file on disk because cv2.imread / PaddleOCR expect paths.
    """
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    file_storage.save(tmp.name)
    tmp.close()
    return tmp.name


# ---------------------------------------------------------------------------
# Core processing: image  →  list of region results
# ---------------------------------------------------------------------------

def _process_image(image_path: str, original_filename: str = "") -> dict:
    temp_files_to_cleanup = []
    
    # Trích xuất Part No từ tên file (ví dụ: QC7-4047-000)
    part_no = ""
    if original_filename:
        # Pattern cho mã linh kiện (VD: QC7-4047-000)
        match = re.search(r'([A-Z0-9]{3}-[A-Z0-9]{4}-[A-Z0-9]{3})', original_filename.upper())
        if match:
            part_no = match.group(1)
            print(f"📌 Extracted Part No from filename: {part_no}")

    try:
        extractor = ImageRegionExtractor(image_path)
        regions = extractor.detect_regions(min_width=30, min_height=30)

        if not regions:
            return {"success": True, "total_regions": 0, "results": []}

        regions.sort(key=lambda r: (r[1], r[0]))

        padding = 10
        source_img = cv2.imread(image_path, cv2.IMREAD_COLOR)
        h_img, w_img = source_img.shape[:2]

        results: list[dict] = []

        for idx, (x, y, w, h) in enumerate(regions, start=1):
            region_result: dict = {"region_id": idx}

            try:
                x_pad  = max(0, x - padding)
                y_pad  = max(0, y - padding)
                x_max  = min(w_img, x + w + padding)
                y_max  = min(h_img, y + h + padding * 3) 
                roi    = source_img[y_pad:y_max, x_pad:x_max].copy()

                if h > w * 1.2:
                    roi = cv2.rotate(roi, cv2.ROTATE_90_CLOCKWISE)

                region_result["cropped_image"] = _img_to_base64(roi)

                # ✅ Crop image
                tmp_crop = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                cv2.imwrite(tmp_crop.name, roi)
                tmp_crop.close()
                temp_files_to_cleanup.append(tmp_crop.name)
                # print(f"  📁 Created: {tmp_crop.name}")

                # ✅ Preprocess
                preprocessed_path, preprocessed_binary, _ = preprocess_image_region(
                    tmp_crop.name, scale=5
                )
                
                if preprocessed_path:
                    temp_files_to_cleanup.append(preprocessed_path)

                # Loại bỏ việc gửi ảnh preprocessed để giảm dung lượng JSON
                # region_result["preprocessed_image"] = _img_gray_to_base64(preprocessed_binary) if preprocessed_binary is not None else ""
                region_result["preprocessed_image"] = "" 

                # OCR
                dimension_text = read_and_process_text(tmp_crop.name)

                if dimension_text and dimension_text.strip():
                    region_result["dimension"] = dimension_text.strip()
                    
                    # --- VALIDATION VỚI DATABASE ---
                    if part_no:
                        val_status, val_msg, excel_data = validator.validate(part_no, dimension_text)
                        print(f"🔍 Region {idx}: OCR='{dimension_text}' | DB_Val='{excel_data['new_tolerance'] if excel_data else 'N/A'}' | Status={val_status}")
                        region_result["status"] = val_status
                        region_result["validation_message"] = val_msg
                        if excel_data:
                            # Đảm bảo dữ liệu từ Excel là chuỗi để an toàn khi JSON serialize (tránh lỗi NaN/Infinity)
                            region_result["excel_data"] = {
                                "spec": str(excel_data.get("spec", "")),
                                "tolerance": str(excel_data.get("tolerance", "")),
                                "new_tolerance": str(excel_data.get("new_tolerance", ""))
                            }
                    else:
                        region_result["status"] = "success"
                else:
                    region_result["dimension"] = ""
                    region_result["status"]    = "no_text_found"

            except Exception as exc:
                traceback.print_exc()
                region_result.setdefault("cropped_image", "")
                region_result.setdefault("preprocessed_image", "")
                region_result["dimension"] = ""
                region_result["status"]    = "error"
                region_result["error"]     = str(exc)

            results.append(region_result)

        return {
            "success":       True,
            "total_regions": len(regions),
            "results":       results,
        }
    
    finally:
        # ✅ CLEANUP CHẮC CHẮN
        print(f"\n🧹 Cleaning up {len(temp_files_to_cleanup)} temp files...")
        deleted_count = 0
        for temp_file in temp_files_to_cleanup:
            if _safe_remove(temp_file):
                deleted_count += 1
                print(f"  ✓ Deleted: {temp_file}")
        
        print(f"✅ Successfully deleted {deleted_count}/{len(temp_files_to_cleanup)} files")
        
        # Force garbage collection
        gc.collect()




def _safe_remove(path: str) -> bool:
    """Delete a file, return True if successful."""
    try:
        if os.path.exists(path):
            os.unlink(path)
            return True
        return False
    except OSError as e:
        print(f"  ⚠️  Failed to delete {path}: {e}")
        return False


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------

@app.route("/api/ocr-dimensions/excel", methods=["POST"])
def ocr_dimensions_excel():
    """
    POST /api/ocr-dimensions/excel
    Content-Type: multipart/form-data
    Field: "image" → PNG file
    
    Returns: Excel file download
    """
    if "image" not in request.files:
        return jsonify({"success": False, "error": "No 'image' field in request"}), 400

    file = request.files["image"]
    if file.filename == "":
        return jsonify({"success": False, "error": "Empty filename"}), 400

    image_path: str | None = None
    try:
        image_path = _save_uploaded_image(file)

        test = cv2.imread(image_path)
        if test is None:
            return jsonify({"success": False, "error": "Unable to decode uploaded image"}), 400

        # Process image
        payload = _process_image(image_path, file.filename)
        
        if not payload.get('success'):
            return jsonify({"success": False, "error": "Processing failed"}), 500

        # Create Excel
        excel_file = create_excel_report(payload)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"OCR_Results_{timestamp}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as exc:
        traceback.print_exc()
        from process_drawing import reset_ocr_instance
        reset_ocr_instance()
        return jsonify({"success": False, "error": str(exc)}), 500

    finally:
        if image_path:
            _safe_remove(image_path)
        gc.collect()


# ============================================================================
# API ENDPOINT - JSON + EXCEL OPTION
# ============================================================================

@app.route("/api/ocr-dimensions", methods=["POST"])
def ocr_dimensions():
    """
    POST /api/ocr-dimensions
    Content-Type: multipart/form-data
    Fields:
        - "image": PNG file (required)
        - "format": "json" or "excel" (optional, default: "json")
    """
    if "image" not in request.files:
        return jsonify({"success": False, "error": "No 'image' field in request"}), 400

    file = request.files["image"]
    if file.filename == "":
        return jsonify({"success": False, "error": "Empty filename"}), 400

    # Check format preference
    output_format = request.form.get('format', 'json').lower()

    image_path: str | None = None
    try:
        image_path = _save_uploaded_image(file)

        test = cv2.imread(image_path)
        if test is None:
            return jsonify({"success": False, "error": "Unable to decode uploaded image"}), 400

        payload = _process_image(image_path, file.filename)
        
        # Return Excel if requested
        if output_format == 'excel':
            excel_file = create_excel_report(payload)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"OCR_Results_{timestamp}.xlsx"
            
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        # Default: return JSON
        return jsonify(payload), 200

    except Exception as exc:
        traceback.print_exc()
        from process_drawing import reset_ocr_instance
        reset_ocr_instance()
        return jsonify({"success": False, "error": str(exc)}), 500

    finally:
        if image_path:
            _safe_remove(image_path)
        gc.collect()

def create_excel_report(ocr_results: dict) -> BytesIO:
    """
    Tạo file Excel từ kết quả OCR.
    
    Args:
        ocr_results: Dict chứa kết quả từ _process_image()
        
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"
    
    # === STYLING ===
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # === HEADER SECTION ===
    ws.merge_cells('A1:E1')
    ws['A1'] = "OCR DIMENSION EXTRACTION REPORT"
    ws['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws['A1'].alignment = center_align
    
    ws['A2'] = "Generated:"
    ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A2'].font = Font(bold=True)
    
    ws['A3'] = "Total Regions:"
    ws['B3'] = ocr_results.get('total_regions', 0)
    ws['A3'].font = Font(bold=True)
    
    success_count = sum(1 for r in ocr_results.get('results', []) if r.get('status') == 'success')
    ws['A4'] = "Successful:"
    ws['B4'] = success_count
    ws['A4'].font = Font(bold=True)
    ws['B4'].fill = success_fill
    
    # === TABLE HEADERS ===
    headers = ['Region ID', 'Status', 'Dimension Result', 'Notes', 'Details']
    header_row = 6
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # === DATA ROWS ===
    current_row = header_row + 1
    
    for result in ocr_results.get('results', []):
        region_id = result.get('region_id', '-')
        status = result.get('status', 'unknown')
        dimension = result.get('dimension', '')
        error = result.get('error', '')
        
        # Region ID
        cell = ws.cell(row=current_row, column=1, value=region_id)
        cell.alignment = center_align
        cell.border = thin_border
        
        # Status
        cell = ws.cell(row=current_row, column=2, value=status.upper())
        cell.alignment = center_align
        cell.border = thin_border
        
        if status == 'success':
            cell.fill = success_fill
            cell.font = Font(bold=True, color="006100")
        elif status == 'error':
            cell.fill = error_fill
            cell.font = Font(bold=True, color="9C0006")
        elif status == 'no_text_found':
            cell.fill = warning_fill
            cell.font = Font(bold=True, color="9C6500")
        
        # Dimension Result
        cell = ws.cell(row=current_row, column=3, value=dimension if dimension else "N/A")
        cell.alignment = left_align
        cell.border = thin_border
        if dimension:
            cell.font = Font(bold=True, size=11)
        
        # Notes
        notes = ""
        if status == 'error':
            notes = f"Error: {error}"
        elif status == 'no_text_found':
            notes = "No text detected in this region"
        elif status == 'success' and dimension:
            # Phân tích các thành phần
            has_ref = '(' in dimension and ')' in dimension
            has_tolerance = '/' in dimension
            has_general_tolerance = '<' in dimension
            
            components = []
            if has_ref:
                components.append("Reference")
            if has_tolerance:
                components.append("Deviation")
            if has_general_tolerance:
                components.append("Tolerance")
            
            notes = f"Contains: {', '.join(components)}" if components else "Valid"
        
        cell = ws.cell(row=current_row, column=4, value=notes)
        cell.alignment = left_align
        cell.border = thin_border
        
        # Details (optional - có thể thêm thông tin về confidence, vị trí, etc.)
        details = f"Region #{region_id}"
        cell = ws.cell(row=current_row, column=5, value=details)
        cell.alignment = center_align
        cell.border = thin_border
        
        current_row += 1
    
    # === SUMMARY SECTION ===
    summary_row = current_row + 2
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws[f'A{summary_row}'] = "SUMMARY STATISTICS"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12, color="4472C4")
    ws[f'A{summary_row}'].alignment = center_align
    
    summary_row += 1
    
    total = len(ocr_results.get('results', []))
    success = sum(1 for r in ocr_results.get('results', []) if r.get('status') == 'success')
    no_text = sum(1 for r in ocr_results.get('results', []) if r.get('status') == 'no_text_found')
    errors = sum(1 for r in ocr_results.get('results', []) if r.get('status') == 'error')
    
    ws[f'A{summary_row}'] = "Total Regions:"
    ws[f'B{summary_row}'] = total
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "✓ Success:"
    ws[f'B{summary_row}'] = success
    ws[f'A{summary_row}'].font = Font(bold=True, color="006100")
    ws[f'B{summary_row}'].fill = success_fill
    
    summary_row += 1
    ws[f'A{summary_row}'] = "⚠ No Text:"
    ws[f'B{summary_row}'] = no_text
    ws[f'A{summary_row}'].font = Font(bold=True, color="9C6500")
    ws[f'B{summary_row}'].fill = warning_fill
    
    summary_row += 1
    ws[f'A{summary_row}'] = "✗ Errors:"
    ws[f'B{summary_row}'] = errors
    ws[f'A{summary_row}'].font = Font(bold=True, color="9C0006")
    ws[f'B{summary_row}'].fill = error_fill
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Success Rate:"
    success_rate = (success / total * 100) if total > 0 else 0
    ws[f'B{summary_row}'] = f"{success_rate:.1f}%"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].font = Font(bold=True, size=11)
    
    # === COLUMN WIDTHS ===
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    
    # === SAVE TO MEMORY ===
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
# ---------------------------------------------------------------------------
# Health-check (optional, useful for monitoring)
# ---------------------------------------------------------------------------

@app.route("/api/health", methods=["GET"])
def health():
    """Quick liveness probe — also warms up the OCR model on first call."""
    try:
        _ = get_ocr_instance()          # lazy-init; no-op after first call
        return jsonify({"status": "ok", "message": f"{OCR_TYPE} OCR ready"}), 200
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


# ---------------------------------------------------------------------------
# Entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Pre-warm OCR so the first real request is fast
    print(f"🔄 Warming up {OCR_TYPE} OCR …")
    _ = get_ocr_instance()
    print(f"✅ {OCR_TYPE} OCR ready.")

    app.run(host="0.0.0.0", port=8080, debug=False)
